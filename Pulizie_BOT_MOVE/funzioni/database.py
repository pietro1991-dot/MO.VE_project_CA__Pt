"""
Database manager per il bot delle pulizie
TUTTO BASATO SU EXCEL - No SQLite
"""

import os
import shutil
import time
import logging
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Tuple
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from filelock import FileLock

logger = logging.getLogger(__name__)

# Path Excel files - Database CONDIVISO (un livello sopra il bot)
EXCEL_DIR = os.path.join(os.path.dirname(__file__), '..', '..', 'Database')
EXCEL_APPARTAMENTI_PATH = os.path.join(EXCEL_DIR, 'appartamenti.xlsx')
EXCEL_USERS_PATH = os.path.join(EXCEL_DIR, 'users.xlsx')
EXCEL_TURNI_PATH = os.path.join(EXCEL_DIR, 'turni.xlsx')
EXCEL_RICHIESTE_PATH = os.path.join(EXCEL_DIR, 'richieste_prodotti.xlsx')
EXCEL_MATERIALI_PATH = os.path.join(EXCEL_DIR, 'materiali_pulizie_appartamenti.xlsx')

# Cache per performance
_cache_timestamp = 0
_cache_turni = []
_last_request_time = {}  # Rate limiting per richieste

# ==================== UTILITY FUNCTIONS ====================

def backup_excel():
    """Crea backup automatico di tutti i file Excel"""
    try:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_dir = os.path.join(EXCEL_DIR, 'backups')
        os.makedirs(backup_dir, exist_ok=True)
        
        files_backed_up = 0
        for file_path in [EXCEL_USERS_PATH, EXCEL_TURNI_PATH, EXCEL_RICHIESTE_PATH, EXCEL_APPARTAMENTI_PATH]:
            if os.path.exists(file_path):
                filename = os.path.basename(file_path)
                backup_path = os.path.join(backup_dir, f"{filename}.{timestamp}.bak")
                shutil.copy2(file_path, backup_path)
                files_backed_up += 1
                logger.info(f"Backup creato: {backup_path}")
        
        # Pulizia backup vecchi (mantieni solo ultimi 30 giorni)
        cleanup_old_backups(backup_dir, days=30)
        
        return files_backed_up
    except Exception as e:
        logger.error(f"Errore durante backup: {e}")
        return 0

def cleanup_old_backups(backup_dir: str, days: int = 30):
    """Elimina backup più vecchi di N giorni"""
    try:
        cutoff = time.time() - (days * 86400)
        for filename in os.listdir(backup_dir):
            filepath = os.path.join(backup_dir, filename)
            if os.path.isfile(filepath) and os.path.getmtime(filepath) < cutoff:
                os.remove(filepath)
                logger.info(f"Backup vecchio eliminato: {filename}")
    except Exception as e:
        logger.error(f"Errore pulizia backup: {e}")

def sanitize_text(text: str, max_length: int = 100) -> str:
    """Sanitizza input utente rimuovendo caratteri pericolosi"""
    if not text:
        return ""
    
    # Rimuovi spazi extra
    text = text.strip()
    
    # Rimuovi caratteri non validi per filesystem e Excel
    import re
    text = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '', text)
    
    # Limita lunghezza
    text = text[:max_length]
    
    return text

def can_create_request(user_id: int, cooldown_seconds: int = 60) -> bool:
    """Rate limiting per richieste prodotti (evita spam)"""
    now = time.time()
    last_time = _last_request_time.get(user_id, 0)
    
    if now - last_time < cooldown_seconds:
        return False
    
    _last_request_time[user_id] = now
    return True

# ==================== DATABASE INITIALIZATION ====================

def init_database():
    """Inizializza i file Excel necessari"""
    os.makedirs(EXCEL_DIR, exist_ok=True)
    
    # Crea users.xlsx se non esiste
    if not os.path.exists(EXCEL_USERS_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        ws.append(['telegram_id', 'username', 'nome', 'cognome', 'phone', 'created_at'])
        # Formatta header
        for cell in ws[1]:
            cell.font = Font(bold=True)
        wb.save(EXCEL_USERS_PATH)
        print(f"✅ Creato {EXCEL_USERS_PATH}")
    
    # Crea turni.xlsx se non esiste
    if not os.path.exists(EXCEL_TURNI_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Turni"
        ws.append(['id', 'user_telegram_id', 'user_nome', 'user_cognome', 'appartamento_id', 
                   'appartamento_nome', 'data', 'timestamp_ingresso', 'timestamp_uscita', 
                   'ore_lavorate', 'video_ingresso_path', 'video_ingresso_file_id',
                   'video_uscita_path', 'video_uscita_file_id', 'status'])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        wb.save(EXCEL_TURNI_PATH)
        print(f"✅ Creato {EXCEL_TURNI_PATH}")
    
    # Crea richieste_prodotti.xlsx se non esiste
    if not os.path.exists(EXCEL_RICHIESTE_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Richieste"
        # Colonne: id, user_telegram_id, user_nome, appartamento_id, appartamento_nome,
        #          tipo_richiesta (pulizie/appartamento), descrizione_prodotti, 
        #          info_consegna, completato, data_richiesta, data_completamento, message_id
        ws.append(['id', 'user_telegram_id', 'user_nome', 'appartamento_id', 'appartamento_nome',
                   'tipo_richiesta', 'descrizione_prodotti', 'info_consegna',
                   'completato', 'data_richiesta', 'data_completamento', 'message_id'])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        wb.save(EXCEL_RICHIESTE_PATH)
        print(f"✅ Creato {EXCEL_RICHIESTE_PATH}")
    else:
        # Verifica se esiste la nuova struttura, altrimenti aggiorna header
        try:
            wb = openpyxl.load_workbook(EXCEL_RICHIESTE_PATH)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            if 'tipo_richiesta' not in headers:
                logger.info("Aggiornamento struttura richieste_prodotti.xlsx...")
                # Aggiungi nuove colonne se non esistono
                # Backup prima
                backup_excel()
        except:
            pass
    
    print("✅ Database Excel inizializzato correttamente")


# ==================== USERS (Excel) ====================

def register_user(telegram_id: int, username: str, nome: str, cognome: str, phone: str = None) -> bool:
    """Registra un nuovo utente in Excel con file locking e validazione"""
    lock = FileLock(f"{EXCEL_USERS_PATH}.lock", timeout=10)
    
    try:
        # Verifica se esiste già
        if user_exists(telegram_id):
            return False
        
        # Sanitizza input
        nome = sanitize_text(nome, max_length=50)
        cognome = sanitize_text(cognome, max_length=50)
        username = sanitize_text(username or '', max_length=50)
        
        # Valida lunghezza minima
        if len(nome) < 2 or len(cognome) < 2:
            logger.error(f"Nome/cognome troppo corti: {nome} {cognome}")
            return False
        
        with lock:
            wb = openpyxl.load_workbook(EXCEL_USERS_PATH)
            ws = wb.active
            
            ws.append([
                telegram_id,
                username,
                nome,
                cognome,
                phone or '',
                datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ])
            
            wb.save(EXCEL_USERS_PATH)
            wb.close()
            logger.info(f"Utente registrato: {nome} {cognome} (ID: {telegram_id})")
            return True
            
    except FileNotFoundError:
        logger.error(f"File Excel non trovato: {EXCEL_USERS_PATH}")
        return False
    except PermissionError:
        logger.error(f"Permessi insufficienti per scrivere su: {EXCEL_USERS_PATH}")
        return False
    except Exception as e:
        logger.error(f"Errore registrazione utente: {e}", exc_info=True)
        return False

def get_user(telegram_id: int) -> Optional[Dict]:
    """Ottiene info utente da Excel"""
    wb = None
    try:
        wb = openpyxl.load_workbook(EXCEL_USERS_PATH, read_only=True)
        ws = wb.active
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == telegram_id:
                return {
                    'telegram_id': row[0],
                    'username': row[1],
                    'nome': row[2],
                    'cognome': row[3],
                    'phone_number': row[4],
                    'created_at': row[5]
                }
        
        return None
    except Exception as e:
        print(f"❌ Errore get_user: {e}")
        return None
    finally:
        if wb:
            wb.close()

def get_all_users() -> List[Dict]:
    """Ottiene tutti gli utenti da Excel"""
    wb = None
    try:
        wb = openpyxl.load_workbook(EXCEL_USERS_PATH, read_only=True)
        ws = wb.active
        
        users = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                users.append({
                    'telegram_id': row[0],
                    'username': row[1],
                    'nome': row[2],
                    'cognome': row[3],
                    'phone_number': row[4],
                    'created_at': row[5]
                })
        
        return users
    except Exception as e:
        print(f"❌ Errore get_all_users: {e}")
        return []
    finally:
        if wb:
            wb.close()

def user_exists(telegram_id: int) -> bool:
    """Verifica se utente è registrato"""
    return get_user(telegram_id) is not None


# ==================== TURNI (Excel) ====================

def _get_next_turno_id() -> int:
    """Ottiene il prossimo ID turno con file locking per evitare duplicati"""
    lock = FileLock(f"{EXCEL_TURNI_PATH}.lock", timeout=10)
    wb = None
    try:
        with lock:
            wb = openpyxl.load_workbook(EXCEL_TURNI_PATH, read_only=True)
            ws = wb.active
            max_id = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and isinstance(row[0], int):
                    max_id = max(max_id, row[0])
            return max_id + 1
    except:
        return 1
    finally:
        if wb:
            wb.close()

def create_turno(user_id: int, appartamento_id: int, video_path: str, 
                 video_file_id: str, timestamp: datetime) -> int:
    """Crea nuovo turno (ingresso) in Excel con file locking e controllo turno doppio"""
    lock = FileLock(f"{EXCEL_TURNI_PATH}.lock", timeout=10)
    
    try:
        # VERIFICA TURNO GIÀ APERTO (previene turni doppi)
        turno_aperto = get_turno_in_corso(user_id)
        if turno_aperto:
            logger.warning(f"Turno già aperto per user {user_id}: {turno_aperto['appartamento_nome']}")
            raise ValueError(f"Hai già un turno aperto all'appartamento {turno_aperto['appartamento_nome']}")
        
        user = get_user(user_id)
        appartamento = get_appartamento(appartamento_id)
        
        if not user or not appartamento:
            logger.error(f"User o appartamento non trovato: user_id={user_id}, app_id={appartamento_id}")
            return 0
        
        turno_id = _get_next_turno_id()
        
        with lock:
            wb = openpyxl.load_workbook(EXCEL_TURNI_PATH)
            ws = wb.active
            
            ws.append([
                turno_id,                                      # id
                user_id,                                       # user_telegram_id
                user['nome'],                                  # user_nome (CORRETTO)
                user['cognome'],                               # user_cognome (CORRETTO)
                appartamento_id,                               # appartamento_id (CORRETTO)
                appartamento['nome'],                          # appartamento_nome (CORRETTO)
                timestamp.date().strftime('%Y-%m-%d'),         # data
                timestamp.strftime('%Y-%m-%d %H:%M:%S'),       # timestamp_ingresso
                '',                                            # timestamp_uscita
                0,                                             # ore_lavorate
                video_path,                                    # video_ingresso_path
                video_file_id,                                 # video_ingresso_file_id
                '',                                            # video_uscita_path
                '',                                            # video_uscita_file_id
                'in_corso'                                     # status
            ])
            
            wb.save(EXCEL_TURNI_PATH)
            wb.close()
            
            # Invalida cache
            global _cache_timestamp
            _cache_timestamp = 0
            
            logger.info(f"Turno {turno_id} creato: {user['nome']} @ {appartamento['nome']}")
            return turno_id
            
    except ValueError:
        # Errore di validazione (turno doppio)
        raise
    except FileNotFoundError:
        logger.error(f"File Excel non trovato: {EXCEL_TURNI_PATH}")
        return 0
    except PermissionError:
        logger.error(f"Permessi insufficienti per scrivere su: {EXCEL_TURNI_PATH}")
        return 0
    except Exception as e:
        logger.error(f"Errore create_turno: {e}", exc_info=True)
        return 0

def get_turno_in_corso(user_id: int) -> Optional[Dict]:
    """Ottiene turno in corso per utente da Excel"""
    wb = None
    try:
        wb = openpyxl.load_workbook(EXCEL_TURNI_PATH, read_only=True)
        ws = wb.active
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] == user_id and row[14] == 'in_corso':  # status = 'in_corso' (colonna 14, non 15!)
                return {
                    'id': row[0],
                    'user_telegram_id': row[1],
                    'nome': row[2],                          # CORRETTO: user_nome
                    'cognome': row[3],                       # CORRETTO: user_cognome
                    'appartamento_id': row[4],               # CORRETTO: appartamento_id
                    'appartamento_nome': row[5],             # CORRETTO: appartamento_nome
                    'data': row[6],
                    'timestamp_ingresso': row[7],
                    'timestamp_uscita': row[8],
                    'ore_lavorate': row[9],
                    'video_ingresso_path': row[10],
                    'video_ingresso_file_id': row[11],
                    'video_uscita_path': row[12],
                    'video_uscita_file_id': row[13],
                    'status': row[14],
                    'indirizzo': ''  # Non serve, ma per compatibilità
                }
        
        return None
    except Exception as e:
        print(f"❌ Errore get_turno_in_corso: {e}")
        return None
    finally:
        if wb:
            wb.close()

def complete_turno(turno_id: int, video_path: str, video_file_id: str, timestamp: datetime):
    """Completa turno (uscita) e calcola ore lavorate in Excel con file locking"""
    lock = FileLock(f"{EXCEL_TURNI_PATH}.lock", timeout=10)
    
    try:
        with lock:
            wb = openpyxl.load_workbook(EXCEL_TURNI_PATH)
            ws = wb.active
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if row[0].value == turno_id:
                    # Calcola ore
                    ts_ingresso_str = row[7].value  # timestamp_ingresso (colonna 7, non 8!)
                    if isinstance(ts_ingresso_str, str):
                        timestamp_ingresso = datetime.strptime(ts_ingresso_str, '%Y-%m-%d %H:%M:%S')
                    else:
                        timestamp_ingresso = ts_ingresso_str
                    
                    ore_lavorate = (timestamp - timestamp_ingresso).total_seconds() / 3600
                    
                    # Valida ore (max 24 ore)
                    if ore_lavorate > 24:
                        logger.warning(f"Turno {turno_id}: ore lavorate sospette ({ore_lavorate:.2f}h)")
                    
                    # Aggiorna riga (correzione indici: le colonne Excel sono 1-indexed!)
                    ws.cell(row_idx, 9, timestamp.strftime('%Y-%m-%d %H:%M:%S'))   # timestamp_uscita (col 9)
                    ws.cell(row_idx, 10, round(ore_lavorate, 2))                    # ore_lavorate (col 10)
                    ws.cell(row_idx, 13, video_path)                                # video_uscita_path (col 13)
                    ws.cell(row_idx, 14, video_file_id)                             # video_uscita_file_id (col 14)
                    ws.cell(row_idx, 15, 'completato')                              # status (col 15)
                    
                    wb.save(EXCEL_TURNI_PATH)
                    wb.close()
                    
                    # Invalida cache
                    global _cache_timestamp
                    _cache_timestamp = 0
                    
                    logger.info(f"Turno {turno_id} completato: {ore_lavorate:.2f}h")
                    return ore_lavorate
            
            wb.close()
            logger.warning(f"Turno {turno_id} non trovato per completamento")
            return None
            
    except FileNotFoundError:
        logger.error(f"File Excel non trovato: {EXCEL_TURNI_PATH}")
        return None
    except PermissionError:
        logger.error(f"Permessi insufficienti per scrivere su: {EXCEL_TURNI_PATH}")
        return None
    except Exception as e:
        logger.error(f"Errore complete_turno: {e}", exc_info=True)
        return None

def get_turni_by_date(data: datetime.date) -> List[Dict]:
    """Ottiene tutti i turni di una data da Excel"""
    try:
        wb = openpyxl.load_workbook(EXCEL_TURNI_PATH, read_only=True)
        ws = wb.active
        
        data_str = data.strftime('%Y-%m-%d')
        turni = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[6] == data_str:  # data è colonna 6, non 7!
                turni.append({
                    'id': row[0],
                    'user_telegram_id': row[1],
                    'nome': row[2],           # CORRETTO
                    'cognome': row[3],        # CORRETTO
                    'appartamento_nome': row[5],  # CORRETTO
                    'data': row[6],           # CORRETTO
                    'timestamp_ingresso': row[7],  # CORRETTO
                    'timestamp_uscita': row[8],    # CORRETTO
                    'ore_lavorate': row[9],   # CORRETTO
                    'status': row[14]         # CORRETTO
                })
        
        wb.close()
        return turni
    except Exception as e:
        print(f"❌ Errore get_turni_by_date: {e}")
        return []

def get_turni_by_user(user_id: int, data_inizio: datetime.date = None, 
                      data_fine: datetime.date = None) -> List[Dict]:
    """Ottiene turni di un utente in un periodo da Excel"""
    wb = None
    try:
        wb = openpyxl.load_workbook(EXCEL_TURNI_PATH, read_only=True)
        ws = wb.active
        
        turni = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] == user_id:
                # Filtra per date se specificate
                data_turno_str = row[6]  # CORRETTO
                if data_inizio or data_fine:
                    data_turno = datetime.strptime(data_turno_str, '%Y-%m-%d').date()
                    if data_inizio and data_turno < data_inizio:
                        continue
                    if data_fine and data_turno > data_fine:
                        continue
                
                turni.append({
                    'id': row[0],
                    'appartamento_nome': row[5],   # CORRETTO
                    'data': row[6],                # CORRETTO
                    'timestamp_ingresso': row[7],  # CORRETTO
                    'timestamp_uscita': row[8],    # CORRETTO
                    'ore_lavorate': row[9],        # CORRETTO
                    'status': row[14]              # CORRETTO
                })
        
        return turni
    except Exception as e:
        print(f"❌ Errore get_turni_by_user: {e}")
        return []
    finally:
        if wb:
            wb.close()

def get_all_turni_in_corso() -> List[Dict]:
    """Ottiene tutti i turni in corso"""
    wb = None
    try:
        wb = openpyxl.load_workbook(EXCEL_TURNI_PATH, read_only=True)
        ws = wb.active
        
        turni = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[14] == 'in_corso':  # status
                turni.append({
                    'id': row[0],
                    'user_id': row[1],
                    'nome': row[2],
                    'cognome': row[3],
                    'appartamento_id': row[4],
                    'appartamento_nome': row[5],
                    'data': row[6],
                    'timestamp_ingresso': row[7],
                    'video_ingresso': row[10]
                })
        
        return turni
    except Exception as e:
        print(f"❌ Errore get_all_turni_in_corso: {e}")
        return []
    finally:
        if wb:
            wb.close()

def get_all_turni_completati(limit: int = 50) -> List[Dict]:
    """Ottiene tutti i turni completati (ultimi N)"""
    wb = None
    try:
        wb = openpyxl.load_workbook(EXCEL_TURNI_PATH, read_only=True)
        ws = wb.active
        
        turni = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[14] == 'completato':  # status
                turni.append({
                    'id': row[0],
                    'user_id': row[1],
                    'nome': row[2],
                    'cognome': row[3],
                    'appartamento_id': row[4],
                    'appartamento_nome': row[5],
                    'data': row[6],
                    'timestamp_ingresso': row[7],
                    'timestamp_uscita': row[8],
                    'ore_lavorate': row[9],
                    'video_ingresso': row[10],
                    'video_uscita': row[12]  # Corretto: video_uscita_path è colonna 12
                })
        
        # Ordina per data decrescente e limita
        turni.sort(key=lambda x: x['timestamp_uscita'] or '', reverse=True)
        return turni[:limit]
    except Exception as e:
        print(f"❌ Errore get_all_turni_completati: {e}")
        return []
    finally:
        if wb:
            wb.close()

def get_turni_completati_oggi() -> List[Dict]:
    """Ottiene tutti i turni completati oggi"""
    wb = None
    try:
        oggi = datetime.now().strftime('%Y-%m-%d')
        wb = openpyxl.load_workbook(EXCEL_TURNI_PATH, read_only=True)
        ws = wb.active
        
        turni = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[14] == 'completato' and row[6] == oggi:  # status e data
                turni.append({
                    'id': row[0],
                    'user_id': row[1],
                    'nome': row[2],
                    'cognome': row[3],
                    'appartamento_id': row[4],
                    'appartamento_nome': row[5],
                    'data': row[6],
                    'timestamp_ingresso': row[7],
                    'timestamp_uscita': row[8],
                    'ore_lavorate': row[9],
                    'video_ingresso': row[10],  # video_ingresso_path
                    'video_uscita': row[12]     # video_uscita_path (corretto da 11 a 12)
                })
        
        # Ordina per timestamp uscita
        turni.sort(key=lambda x: x['timestamp_uscita'] or '', reverse=True)
        return turni
    except Exception as e:
        print(f"❌ Errore get_turni_completati_oggi: {e}")
        return []
    finally:
        if wb:
            wb.close()

def esporta_turni_excel(turni: List[Dict], titolo: str = "Turni") -> BytesIO:
    """Esporta lista turni in file Excel"""
    from io import BytesIO
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titolo[:31]  # Max 31 caratteri per nome foglio
    
    try:
        # Intestazioni
        headers = ['ID', 'Nome', 'Cognome', 'Appartamento', 'Data', 
                   'Ora Ingresso', 'Ora Uscita', 'Ore Lavorate']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
        
        # Dati
        for row_num, turno in enumerate(turni, 2):
            ws.cell(row=row_num, column=1, value=turno.get('id', ''))
            ws.cell(row=row_num, column=2, value=turno.get('nome', ''))
            ws.cell(row=row_num, column=3, value=turno.get('cognome', ''))
            ws.cell(row=row_num, column=4, value=turno.get('appartamento_nome', ''))
            ws.cell(row=row_num, column=5, value=turno.get('data', ''))
            
            # Formatta ore
            ts_in = turno.get('timestamp_ingresso', '')
            ts_out = turno.get('timestamp_uscita', '')
            if ts_in:
                try:
                    dt = datetime.fromisoformat(ts_in)
                    ts_in = dt.strftime('%H:%M')
                except:
                    pass
            if ts_out:
                try:
                    dt = datetime.fromisoformat(ts_out)
                    ts_out = dt.strftime('%H:%M')
                except:
                    pass
            
            ws.cell(row=row_num, column=6, value=ts_in)
            ws.cell(row=row_num, column=7, value=ts_out)
            ws.cell(row=row_num, column=8, value=turno.get('ore_lavorate', 0) or 0)
        
        # Autofit colonne (approssimativo)
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)
        
        # Salva in BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    finally:
        wb.close()

def get_ore_totali_user(user_id: int, data: datetime.date = None) -> float:
    """Calcola ore totali lavorate per utente in una data da Excel"""
    try:
        turni = get_turni_by_user(user_id, data, data) if data else get_turni_by_user(user_id)
        totale = sum(t.get('ore_lavorate', 0) or 0 for t in turni if t['status'] == 'completato')
        return round(totale, 2)
    except:
        return 0.0


# ==================== RICHIESTE PRODOTTI (Excel + Telegram) ====================

def _get_next_richiesta_id() -> int:
    """Ottiene il prossimo ID richiesta con file locking per evitare duplicati"""
    lock = FileLock(f"{EXCEL_RICHIESTE_PATH}.lock", timeout=10)
    wb = None
    try:
        with lock:
            wb = openpyxl.load_workbook(EXCEL_RICHIESTE_PATH, read_only=True)
            ws = wb.active
            max_id = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and isinstance(row[0], int):
                    max_id = max(max_id, row[0])
            return max_id + 1
    except:
        return 1
    finally:
        if wb:
            wb.close()

def create_richiesta(user_id: int, appartamento_id: int, descrizione: str, 
                     tipo_richiesta: str = 'generico', info_consegna: str = '',
                     turno_id: int = None, message_id: int = None) -> int:
    """Crea nuova richiesta prodotti in Excel con rate limiting e file locking
    
    Args:
        tipo_richiesta: 'pulizie' | 'appartamento' | 'generico'
        info_consegna: Info su luogo e data consegna (opzionale)
    """
    lock = FileLock(f"{EXCEL_RICHIESTE_PATH}.lock", timeout=10)
    
    try:
        # RATE LIMITING (evita spam di richieste)
        if not can_create_request(user_id, cooldown_seconds=30):
            logger.warning(f"Rate limit superato per user {user_id}")
            raise ValueError("⚠️ Aspetta almeno 30 secondi prima di inviare un'altra richiesta")
        
        user = get_user(user_id)
        appartamento = get_appartamento(appartamento_id)
        
        if not user or not appartamento:
            logger.error(f"User o appartamento non trovato: user_id={user_id}, app_id={appartamento_id}")
            return 0
        
        # Sanitizza descrizione
        descrizione = sanitize_text(descrizione, max_length=500)
        if len(descrizione) < 3:
            logger.warning(f"Descrizione troppo corta: '{descrizione}'")
            raise ValueError("⚠️ Descrizione troppo corta (minimo 3 caratteri)")
        
        # Sanitizza info consegna
        info_consegna = sanitize_text(info_consegna, max_length=300)
        
        richiesta_id = _get_next_richiesta_id()
        
        with lock:
            wb = openpyxl.load_workbook(EXCEL_RICHIESTE_PATH)
            ws = wb.active
            
            # Nuova struttura: id, user_telegram_id, user_nome, appartamento_id, appartamento_nome,
            #                  tipo_richiesta, descrizione_prodotti, info_consegna,
            #                  completato, data_richiesta, data_completamento, message_id
            ws.append([
                richiesta_id,
                user_id,
                f"{user['nome']} {user['cognome']}",
                appartamento_id,
                appartamento['nome'],
                tipo_richiesta,
                descrizione,
                info_consegna,
                'NO',  # completato
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                '',  # data_completamento
                message_id or ''  # message_id per aggiornare messaggio Telegram
            ])
            
            wb.save(EXCEL_RICHIESTE_PATH)
            wb.close()
            
            logger.info(f"Richiesta {richiesta_id} creata: {user['nome']} @ {appartamento['nome']}")
            return richiesta_id
            
    except ValueError:
        # Errore di validazione (rate limit o descrizione corta)
        raise
    except FileNotFoundError:
        logger.error(f"File Excel non trovato: {EXCEL_RICHIESTE_PATH}")
        return 0
    except PermissionError:
        logger.error(f"Permessi insufficienti per scrivere su: {EXCEL_RICHIESTE_PATH}")
        return 0
    except Exception as e:
        logger.error(f"Errore create_richiesta: {e}", exc_info=True)
        return 0

def get_richieste_non_completate() -> List[Dict]:
    """Ottiene tutte le richieste non completate da Excel"""
    wb = None
    try:
        wb = openpyxl.load_workbook(EXCEL_RICHIESTE_PATH, read_only=True)
        ws = wb.active
        
        richieste = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Nuova struttura: completato è colonna 8 (indice 8)
            if row[8] == 'NO':  # completato = 'NO'
                richieste.append({
                    'id': row[0],
                    'user_telegram_id': row[1],
                    'user_nome_completo': row[2],
                    'nome': row[2].split()[0] if row[2] else '',
                    'cognome': ' '.join(row[2].split()[1:]) if row[2] else '',
                    'appartamento_id': row[3],
                    'appartamento_nome': row[4],
                    'tipo_richiesta': row[5] if len(row) > 5 else 'generico',
                    'descrizione_prodotti': row[6],
                    'info_consegna': row[7] if len(row) > 7 else '',
                    'completato': False,
                    'data_richiesta': row[9],
                    'message_id': row[11] if len(row) > 11 else None
                })
        
        return richieste
    except Exception as e:
        print(f"❌ Errore get_richieste_non_completate: {e}")
        return []
    finally:
        if wb:
            wb.close()

def complete_richiesta(richiesta_id: int):
    """Segna richiesta come completata in Excel con file locking"""
    lock = FileLock(f"{EXCEL_RICHIESTE_PATH}.lock", timeout=10)
    
    try:
        with lock:
            wb = openpyxl.load_workbook(EXCEL_RICHIESTE_PATH)
            ws = wb.active
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if row[0].value == richiesta_id:
                    ws.cell(row_idx, 9, 'SI')  # completato (colonna 9)
                    ws.cell(row_idx, 11, datetime.now().strftime('%Y-%m-%d %H:%M:%S'))  # data_completamento (colonna 11)
                    wb.save(EXCEL_RICHIESTE_PATH)
                    logger.info(f"Richiesta {richiesta_id} completata")
                    break
            
            wb.close()
            
    except FileNotFoundError:
        logger.error(f"File Excel non trovato: {EXCEL_RICHIESTE_PATH}")
    except PermissionError:
        logger.error(f"Permessi insufficienti per scrivere su: {EXCEL_RICHIESTE_PATH}")
    except Exception as e:
        logger.error(f"Errore complete_richiesta: {e}", exc_info=True)

def delete_richieste_completate():
    """Elimina tutte le richieste completate da Excel con file locking"""
    lock = FileLock(f"{EXCEL_RICHIESTE_PATH}.lock", timeout=10)
    
    try:
        with lock:
            wb = openpyxl.load_workbook(EXCEL_RICHIESTE_PATH)
            ws = wb.active
            
            rows_to_delete = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if row[8].value == 'SI':  # completato = 'SI' (colonna 9, indice 8)
                    rows_to_delete.append(row_idx)
            
            # Elimina dall'ultima riga verso l'alto per evitare problemi di indice
            for row_idx in reversed(rows_to_delete):
                ws.delete_rows(row_idx)
            
            wb.save(EXCEL_RICHIESTE_PATH)
            wb.close()
            
            logger.info(f"Eliminate {len(rows_to_delete)} richieste completate")
            return len(rows_to_delete)
            
    except FileNotFoundError:
        logger.error(f"File Excel non trovato: {EXCEL_RICHIESTE_PATH}")
        return 0
    except PermissionError:
        logger.error(f"Permessi insufficienti per scrivere su: {EXCEL_RICHIESTE_PATH}")
        return 0
    except Exception as e:
        logger.error(f"Errore delete_richieste_completate: {e}", exc_info=True)
        return 0

def get_richiesta(richiesta_id: int) -> Optional[Dict]:
    """Ottiene dettagli richiesta da Excel"""
    wb = None
    try:
        wb = openpyxl.load_workbook(EXCEL_RICHIESTE_PATH, read_only=True)
        ws = wb.active
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == richiesta_id:
                return {
                    'id': row[0],
                    'telegram_id': row[1],
                    'user_nome_completo': row[2],
                    'nome': row[2].split()[0] if row[2] else '',
                    'cognome': ' '.join(row[2].split()[1:]) if row[2] else '',
                    'appartamento_id': row[3],
                    'appartamento_nome': row[4],
                    'tipo_richiesta': row[5] if len(row) > 5 else 'generico',
                    'descrizione_prodotti': row[6],
                    'info_consegna': row[7] if len(row) > 7 else '',
                    'completato': row[8] == 'SI',
                    'data_richiesta': row[9],
                    'message_id': row[11] if len(row) > 11 else None
                }
        
        return None
    except Exception as e:
        print(f"❌ Errore get_richiesta: {e}")
        return None
    finally:
        if wb:
            wb.close()

def update_richiesta_message_id(richiesta_id: int, message_id: int):
    """Aggiorna il message_id Telegram della richiesta per edit successivo con file locking"""
    lock = FileLock(f"{EXCEL_RICHIESTE_PATH}.lock", timeout=10)
    
    try:
        with lock:
            wb = openpyxl.load_workbook(EXCEL_RICHIESTE_PATH)
            ws = wb.active
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if row[0].value == richiesta_id:
                    ws.cell(row_idx, 12, message_id)  # Colonna 12 = message_id
                    wb.save(EXCEL_RICHIESTE_PATH)
                    logger.info(f"Message ID {message_id} salvato per richiesta {richiesta_id}")
                    break
            
            wb.close()
            
    except FileNotFoundError:
        logger.error(f"File Excel non trovato: {EXCEL_RICHIESTE_PATH}")
    except PermissionError:
        logger.error(f"Permessi insufficienti per scrivere su: {EXCEL_RICHIESTE_PATH}")
    except Exception as e:
        logger.error(f"Errore update_richiesta_message_id: {e}", exc_info=True)


# ==================== APPARTAMENTI (da Excel) ====================

def get_appartamento(appartamento_id: int) -> Optional[Dict]:
    """Ottiene info appartamento da Excel"""
    appartamenti = get_all_appartamenti()
    for app in appartamenti:
        if app['id'] == appartamento_id:
            return app
    return None

def get_all_appartamenti() -> List[Dict]:
    """Legge tutti gli appartamenti dall'Excel"""
    if not os.path.exists(EXCEL_APPARTAMENTI_PATH):
        print(f"⚠️ File Excel appartamenti non trovato: {EXCEL_APPARTAMENTI_PATH}")
        return []
    
    wb = None
    try:
        wb = openpyxl.load_workbook(EXCEL_APPARTAMENTI_PATH, read_only=True)
        sheet = wb.active
        
        appartamenti = []
        
        # Leggi righe Excel: A=Gestione, B=Nome, C=Nome OTA, D=Indirizzo, M=Coordinate GPS (colonna 12)
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
            if not row or not row[1]:  # Salta righe vuote (verifica colonna B = nome)
                continue
            
            # Colonna M (indice 12) per coordinate GPS nel formato "lat,lon"
            coordinate = str(row[12]).strip() if len(row) > 12 and row[12] and str(row[12]).strip() not in ['', 'None', 'Vero', 'Falso'] else None
            
            appartamento = {
                'id': row_num,  # Usa numero riga come ID
                'nome': str(row[1]) if len(row) > 1 and row[1] else 'Senza nome',  # Colonna B
                'indirizzo': str(row[3]) if len(row) > 3 and row[3] else '',  # Colonna D
                'coordinate': coordinate,  # Colonna M
                'attivo': str(row[4]) if len(row) > 4 and row[4] else 'Vero'  # Colonna E
            }
            
            appartamenti.append(appartamento)
        
        # NON fare geocoding automatico - troppo lento
        # Il geocoding viene fatto solo quando serve (ricerca GPS)
        
        return appartamenti
        
    except Exception as e:
        print(f"❌ Errore lettura Excel appartamenti: {e}")
        return []
    finally:
        if wb:
            wb.close()


# ==================== MATERIALI PULIZIE E APPARTAMENTO ====================

def get_materiali_pulizie() -> List[str]:
    """Legge i materiali pulizie dal foglio Excel"""
    wb = None
    try:
        if not os.path.exists(EXCEL_MATERIALI_PATH):
            logger.warning(f"File materiali non trovato: {EXCEL_MATERIALI_PATH}")
            return []
        
        wb = openpyxl.load_workbook(EXCEL_MATERIALI_PATH, read_only=True)
        ws = wb['materiali_pulizie']
        
        materiali = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and str(row[0]).strip():
                materiali.append(str(row[0]).strip())
        
        return materiali
    except Exception as e:
        logger.error(f"Errore lettura materiali pulizie: {e}")
        return []
    finally:
        if wb:
            wb.close()


def get_materiali_appartamento() -> List[str]:
    """Legge i materiali appartamento (operazioni mensili) dal foglio Excel"""
    wb = None
    try:
        if not os.path.exists(EXCEL_MATERIALI_PATH):
            logger.warning(f"File materiali non trovato: {EXCEL_MATERIALI_PATH}")
            return []
        
        wb = openpyxl.load_workbook(EXCEL_MATERIALI_PATH, read_only=True)
        ws = wb['materiali_appartamento']
        
        materiali = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and str(row[0]).strip():
                materiali.append(str(row[0]).strip())
        
        return materiali
    except Exception as e:
        logger.error(f"Errore lettura materiali appartamento: {e}")
        return []
    finally:
        if wb:
            wb.close()


# ==================== ACTIVITY LOG (opzionale - per ora skip) ====================

def log_activity(user_id: int, azione: str, dettagli: str = None):
    """Log attività utente - per ora solo print, volendo si può aggiungere log.xlsx"""
    try:
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        print(f"📝 LOG [{timestamp}] User {user_id}: {azione} - {dettagli}")
    except:
        pass


# ==================== REPORT ====================

def get_report_giornaliero(data: datetime.date) -> Dict:
    """Report completo di una giornata"""
    turni = get_turni_by_date(data)
    
    # Calcola statistiche
    ore_totali = sum(t.get('ore_lavorate', 0) or 0 for t in turni if t['status'] == 'completato')
    turni_completati = len([t for t in turni if t['status'] == 'completato'])
    turni_in_corso = len([t for t in turni if t['status'] == 'in_corso'])
    
    # Raggruppa per utente
    utenti = {}
    for t in turni:
        user_key = f"{t['nome']} {t['cognome']}"
        if user_key not in utenti:
            utenti[user_key] = {
                'turni': [],
                'ore_totali': 0
            }
        utenti[user_key]['turni'].append(t)
        if t.get('ore_lavorate'):
            utenti[user_key]['ore_totali'] += t['ore_lavorate']
    
    return {
        'data': data,
        'ore_totali': ore_totali,
        'turni_completati': turni_completati,
        'turni_in_corso': turni_in_corso,
        'utenti': utenti,
        'turni': turni
    }


if __name__ == '__main__':
    # Test inizializzazione
    init_database()
