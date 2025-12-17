"""
Handlers per le funzionalità amministratore
Gestisce pannello admin, report ore, richieste prodotti, archivio video
"""

import logging
from datetime import datetime, timedelta
from io import BytesIO

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ContextTypes
from telegram.error import BadRequest

from . import database as db
from .utils import (
    format_turno_info, format_richiesta_info, format_ora, format_ore,
    format_data, format_data_italiana, get_settimana_corrente, get_mese_corrente
)
from .video_handler import send_video_by_file_id, list_videos_by_date, get_storage_stats
from .config import ADMIN_TELEGRAM_ID, is_admin
from .utils import format_ora

logger = logging.getLogger(__name__)


# ==================== PANNELLO ADMIN ====================

async def cmd_admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Comando /admin - Pannello amministratore"""
    user_id = update.effective_user.id
    
    if not is_admin(user_id):
        await update.message.reply_text("❌ Non hai i permessi per accedere al pannello admin")
        return
    
    keyboard = [
        [InlineKeyboardButton("📋 Turni in corso", callback_data="admin_turni")],
        [InlineKeyboardButton("✅ Turni finiti", callback_data="admin_turni_menu")],
        [InlineKeyboardButton("📦 Richieste prodotti", callback_data="admin_richieste")],
        [InlineKeyboardButton("⏰ Report ore", callback_data="admin_report")],
        [InlineKeyboardButton("📹 Archivio video", callback_data="admin_video")],
        [InlineKeyboardButton("👥 Gestione utenti", callback_data="admin_utenti")],
        [InlineKeyboardButton("📊 Statistiche", callback_data="admin_stats")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    text = "👨‍💼 *PANNELLO AMMINISTRATORE*\n\n"
    text += "Seleziona un'opzione:"
    
    if update.callback_query:
        await update.callback_query.edit_message_text(
            text, reply_markup=reply_markup, parse_mode='Markdown'
        )
    else:
        await update.message.reply_text(
            text, reply_markup=reply_markup, parse_mode='Markdown'
        )


# ==================== TURNI IN CORSO ====================

async def admin_turni_in_corso(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Mostra turni in corso"""
    query = update.callback_query
    is_callback = query is not None
    
    if is_callback:
        await query.answer()
    
    if not is_admin(update.effective_user.id):
        return
    
    # Usa la funzione del database
    turni = db.get_all_turni_in_corso()
    
    if not turni:
        text = "📋 *Turni in corso*\n\n"
        text += "Nessun turno in corso al momento."
    else:
        text = f"📋 *Turni in corso* ({len(turni)})\n\n"
        
        for turno in turni:
            ts_ingresso = datetime.fromisoformat(turno['timestamp_ingresso'])
            durata = datetime.now() - ts_ingresso
            ore_attuali = durata.total_seconds() / 3600
            
            text += f"👤 *{turno['nome']} {turno['cognome']}*\n"
            text += f"🏠 {turno['appartamento_nome']}\n"
            text += f"⏰ Ingresso: {format_ora(ts_ingresso)}\n"
            text += f"⏱️  In corso da: {format_ore(ore_attuali)}\n"
            
            # Alert se turno troppo lungo
            if ore_attuali > 8:
                text += "⚠️ _Turno aperto da più di 8 ore!_\n"
            
            text += "\n"
    
    keyboard = [
        [InlineKeyboardButton("🔄 Aggiorna", callback_data="admin_turni")],
        [InlineKeyboardButton("« Indietro", callback_data="admin_menu")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    if is_callback:
        try:
            await query.edit_message_text(text, reply_markup=reply_markup, parse_mode='Markdown')
        except BadRequest as e:
            if "Message is not modified" not in str(e):
                raise
            await query.answer("✅ Lista già aggiornata")
    else:
        await update.message.reply_text(text, reply_markup=reply_markup, parse_mode='Markdown')


async def admin_turni_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Menu turni completati - scegli oggi/globali/esporta"""
    query = update.callback_query
    is_callback = query is not None
    
    if is_callback:
        try:
            await query.answer()
        except BadRequest:
            pass
    
    if not is_admin(update.effective_user.id):
        return
    
    # Conta turni oggi e globali per mostrare numeri
    turni_oggi = db.get_turni_completati_oggi()
    turni_globali = db.get_all_turni_completati(limit=100)
    
    text = "✅ *TURNI COMPLETATI*\n\n"
    text += f"📅 Turni di oggi: *{len(turni_oggi)}*\n"
    text += f"📊 Turni totali: *{len(turni_globali)}*\n\n"
    text += "Seleziona cosa visualizzare:"
    
    keyboard = [
        [InlineKeyboardButton(f"📅 Turni di Oggi ({len(turni_oggi)})", callback_data="admin_turni_oggi")],
        [InlineKeyboardButton(f"📊 Tutti i Turni ({len(turni_globali)})", callback_data="admin_turni_globali")],
        [InlineKeyboardButton("📥 Esporta Oggi (Excel)", callback_data="admin_export_turni_oggi")],
        [InlineKeyboardButton("📥 Esporta Tutti (Excel)", callback_data="admin_export_turni_globali")],
        [InlineKeyboardButton("« Indietro", callback_data="admin_menu")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    if is_callback:
        try:
            await query.edit_message_text(text, reply_markup=reply_markup, parse_mode='Markdown')
        except BadRequest as e:
            if "Message is not modified" not in str(e) and "Message to edit not found" not in str(e):
                raise
    else:
        await update.message.reply_text(text, reply_markup=reply_markup, parse_mode='Markdown')


async def admin_turni_oggi(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Mostra turni completati oggi"""
    query = update.callback_query
    is_callback = query is not None
    
    if is_callback:
        try:
            await query.answer()
        except BadRequest:
            pass
    
    if not is_admin(update.effective_user.id):
        return
    
    turni = db.get_turni_completati_oggi()
    oggi = datetime.now().strftime('%d/%m/%Y')
    
    if not turni:
        text = f"📅 *Turni completati oggi* ({oggi})\n\n"
        text += "Nessun turno completato oggi."
    else:
        text = f"📅 *Turni completati oggi* ({oggi}) - {len(turni)} turni\n\n"
        ore_totali = 0
        
        for turno in turni:
            ts_ingresso = datetime.fromisoformat(turno['timestamp_ingresso'])
            ts_uscita = datetime.fromisoformat(turno['timestamp_uscita']) if turno['timestamp_uscita'] else None
            ore = turno.get('ore_lavorate', 0) or 0
            ore_totali += ore
            
            text += f"👤 *{turno['nome']} {turno['cognome']}*\n"
            text += f"🏠 {turno['appartamento_nome']}\n"
            text += f"⏰ {format_ora(ts_ingresso)}"
            if ts_uscita:
                text += f" - {format_ora(ts_uscita)}"
            text += f" ({format_ore(ore)})\n"
            text += "─" * 20 + "\n"
            
            if len(text) > 3500:
                text += "\n_...altri turni non mostrati_"
                break
        
        text += f"\n⏱️ *Ore totali oggi: {format_ore(ore_totali)}*"
    
    keyboard = [
        [InlineKeyboardButton("🔄 Aggiorna", callback_data="admin_turni_oggi")],
        [InlineKeyboardButton("📥 Esporta Excel", callback_data="admin_export_turni_oggi")],
        [InlineKeyboardButton("« Indietro", callback_data="admin_turni_menu")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    if is_callback:
        try:
            await query.edit_message_text(text, reply_markup=reply_markup, parse_mode='Markdown')
        except BadRequest as e:
            if "Message is not modified" not in str(e) and "Message to edit not found" not in str(e):
                raise
    else:
        await update.message.reply_text(text, reply_markup=reply_markup, parse_mode='Markdown')


async def admin_turni_globali(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Mostra tutti i turni completati (ultimi 50)"""
    query = update.callback_query
    is_callback = query is not None
    
    if is_callback:
        try:
            await query.answer()
        except BadRequest:
            pass
    
    if not is_admin(update.effective_user.id):
        return
    
    turni = db.get_all_turni_completati(limit=50)
    
    if not turni:
        text = "📊 *Tutti i turni completati*\n\n"
        text += "Nessun turno completato."
    else:
        text = f"📊 *Tutti i turni completati* (ultimi {len(turni)})\n\n"
        
        for turno in turni:
            ts_ingresso = datetime.fromisoformat(turno['timestamp_ingresso'])
            ts_uscita = datetime.fromisoformat(turno['timestamp_uscita']) if turno['timestamp_uscita'] else None
            ore = turno.get('ore_lavorate', 0) or 0
            
            text += f"👤 *{turno['nome']} {turno['cognome']}*\n"
            text += f"🏠 {turno['appartamento_nome']}\n"
            text += f"📅 {turno['data']} | ⏰ {format_ora(ts_ingresso)}"
            if ts_uscita:
                text += f" - {format_ora(ts_uscita)}"
            text += f" ({format_ore(ore)})\n"
            text += "─" * 20 + "\n"
            
            if len(text) > 3500:
                text += "\n_...altri turni non mostrati_"
                break
    
    keyboard = [
        [InlineKeyboardButton("🔄 Aggiorna", callback_data="admin_turni_globali")],
        [InlineKeyboardButton("📥 Esporta Excel", callback_data="admin_export_turni_globali")],
        [InlineKeyboardButton("« Indietro", callback_data="admin_turni_menu")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    if is_callback:
        try:
            await query.edit_message_text(text, reply_markup=reply_markup, parse_mode='Markdown')
        except BadRequest as e:
            if "Message is not modified" not in str(e) and "Message to edit not found" not in str(e):
                raise
    else:
        await update.message.reply_text(text, reply_markup=reply_markup, parse_mode='Markdown')


async def admin_export_turni(update: Update, context: ContextTypes.DEFAULT_TYPE, tipo: str = 'oggi'):
    """Esporta turni in Excel"""
    query = update.callback_query
    
    try:
        await query.answer("📥 Generazione file Excel...")
    except BadRequest:
        pass
    
    if not is_admin(update.effective_user.id):
        return
    
    if tipo == 'oggi':
        turni = db.get_turni_completati_oggi()
        oggi = datetime.now().strftime('%Y%m%d')
        filename = f"turni_oggi_{oggi}.xlsx"
        titolo = f"Turni {datetime.now().strftime('%d/%m/%Y')}"
    else:
        turni = db.get_all_turni_completati(limit=500)
        filename = f"turni_globali_{datetime.now().strftime('%Y%m%d')}.xlsx"
        titolo = "Tutti i Turni"
    
    if not turni:
        await query.message.reply_text("❌ Nessun turno da esportare.")
        return
    
    # Genera Excel
    excel_file = db.esporta_turni_excel(turni, titolo)
    
    # Invia file
    await query.message.reply_document(
        document=excel_file,
        filename=filename,
        caption=f"📥 *{titolo}*\n\n{len(turni)} turni esportati.",
        parse_mode='Markdown'
    )


async def admin_turni_finiti(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Alias per retrocompatibilità - reindirizza a menu turni"""
    await admin_turni_menu(update, context)


# ==================== RICHIESTE PRODOTTI ====================

async def mostra_richieste_in_sospeso(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler per pulsante 'Richieste in Sospeso' - accessibile solo da admin"""
    user_id = update.effective_user.id
    
    if not is_admin(user_id):
        await update.message.reply_text("❌ Non hai i permessi per questa funzione")
        return
    
    richieste = db.get_richieste_non_completate()
    
    keyboard = []
    
    if not richieste:
        text = "📋 *RICHIESTE IN SOSPESO*\n\n"
        text += "✅ Nessuna richiesta in sospeso!\n\n"
        text += "_Tutte le richieste sono state completate._"
    else:
        text = f"📋 *RICHIESTE IN SOSPESO* ({len(richieste)})\n\n"
        
        for r in richieste:
            dt = datetime.fromisoformat(r['data_richiesta'])
            tipo = r.get('tipo_richiesta', 'generico')
            info_consegna = r.get('info_consegna', '')
            
            # Emoji in base al tipo
            if tipo == 'pulizie':
                emoji = "🧹"
                tipo_text = "Materiale Pulizie"
            elif tipo == 'appartamento':
                emoji = "🏠"
                tipo_text = "Manca Appartamento"
            else:
                emoji = "📦"
                tipo_text = "Richiesta"
            
            text += f"{emoji} *{tipo_text}*\n"
            text += f"👤 {r['nome']} {r['cognome']}\n"
            text += f"🏠 {r['appartamento_nome']}\n"
            text += f"📦 {r['descrizione_prodotti']}\n"
            
            if info_consegna:
                text += f"📝 Info: {info_consegna}\n"
            
            text += f"📅 {format_ora(dt)}\n"
            text += "─" * 20 + "\n\n"
            
            # Bottone per completare (mostra tipo + descrizione abbreviata)
            label = f"✅ {tipo_text[:10]} - {r['descrizione_prodotti'][:25]}..."
            keyboard.append([
                InlineKeyboardButton(label, callback_data=f"completa_{r['id']}")
            ])
    
    keyboard.append([InlineKeyboardButton("🔄 Aggiorna lista", callback_data="richieste_aggiorna")])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.message.reply_text(text, reply_markup=reply_markup, parse_mode='Markdown')


async def aggiorna_richieste_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Callback per aggiornare lista richieste in sospeso"""
    query = update.callback_query
    await query.answer("🔄 Aggiornamento...")
    
    if not is_admin(update.effective_user.id):
        return
    
    richieste = db.get_richieste_non_completate()
    
    keyboard = []
    
    if not richieste:
        text = "📋 *RICHIESTE IN SOSPESO*\n\n"
        text += "✅ Nessuna richiesta in sospeso!\n\n"
        text += "_Tutte le richieste sono state completate._"
    else:
        text = f"📋 *RICHIESTE IN SOSPESO* ({len(richieste)})\n\n"
        
        for r in richieste:
            dt = datetime.fromisoformat(r['data_richiesta'])
            tipo = r.get('tipo_richiesta', 'generico')
            info_consegna = r.get('info_consegna', '')
            
            if tipo == 'pulizie':
                emoji = "🧹"
                tipo_text = "Materiale Pulizie"
            elif tipo == 'appartamento':
                emoji = "🏠"
                tipo_text = "Manca Appartamento"
            else:
                emoji = "📦"
                tipo_text = "Richiesta"
            
            text += f"{emoji} *{tipo_text}*\n"
            text += f"👤 {r['nome']} {r['cognome']}\n"
            text += f"🏠 {r['appartamento_nome']}\n"
            text += f"📦 {r['descrizione_prodotti']}\n"
            
            if info_consegna:
                text += f"📝 Info: {info_consegna}\n"
            
            text += f"📅 {format_ora(dt)}\n"
            text += "─" * 20 + "\n\n"
            
            label = f"✅ {tipo_text[:10]} - {r['descrizione_prodotti'][:25]}..."
            keyboard.append([
                InlineKeyboardButton(label, callback_data=f"completa_{r['id']}")
            ])
    
    keyboard.append([InlineKeyboardButton("🔄 Aggiorna lista", callback_data="richieste_aggiorna")])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    try:
        await query.edit_message_text(text, reply_markup=reply_markup, parse_mode='Markdown')
    except BadRequest as e:
        error_msg = str(e)
        if "Message is not modified" in error_msg:
            await query.answer("✅ Lista già aggiornata")
        elif "Message to edit not found" in error_msg:
            logger.error(f"Errore aggiornamento messaggio: {e}")
            await query.message.reply_text(text, reply_markup=reply_markup, parse_mode='Markdown')
        else:
            raise


async def admin_richieste_prodotti(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Mostra richieste prodotti con bottoni per completarle"""
    query = update.callback_query
    try:
        await query.answer()
    except BadRequest:
        pass
    
    if not is_admin(update.effective_user.id):
        return
    
    richieste = db.get_richieste_non_completate()
    
    keyboard = []
    
    if not richieste:
        text = "📦 *PRODOTTI MANCANTI*\n\n"
        text += "✅ Nessuna richiesta in sospeso!"
    else:
        text = f"📦 *PRODOTTI MANCANTI* ({len(richieste)})\n\n"
        
        for r in richieste:
            dt = datetime.fromisoformat(r['data_richiesta'])
            
            text += f"⬜ *{r['appartamento_nome']}*\n"
            text += f"   └ {r['descrizione_prodotti']}\n"
            text += f"   _({r['nome']} {r['cognome']}, {format_ora(dt)})_\n\n"
            
            # Bottone per completare
            keyboard.append([
                InlineKeyboardButton(
                    f"✅ {r['descrizione_prodotti'][:35]}...",
                    callback_data=f"completa_{r['id']}"
                )
            ])
    
    keyboard.append([InlineKeyboardButton("🔄 Aggiorna", callback_data="admin_richieste")])
    keyboard.append([InlineKeyboardButton("🗑️ Rimuovi completati", callback_data="admin_pulisci_richieste")])
    keyboard.append([InlineKeyboardButton("« Indietro", callback_data="admin_menu")])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    try:
        await query.edit_message_text(text, reply_markup=reply_markup, parse_mode='Markdown')
    except BadRequest as e:
        error_msg = str(e)
        if "Message is not modified" in error_msg:
            pass  # Ignora messaggio identico
        elif "Message to edit not found" in error_msg:
            logger.error(f"Errore aggiornamento messaggio: {e}")
            await query.message.reply_text(text, reply_markup=reply_markup, parse_mode='Markdown')
        else:
            raise


async def completa_richiesta(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Segna richiesta come completata e aggiorna messaggio con strikethrough"""
    query = update.callback_query
    
    if not is_admin(update.effective_user.id):
        await query.answer("❌ Non autorizzato")
        return
    
    # Estrai ID richiesta
    richiesta_id = int(query.data.split('_')[1])
    
    # Ottieni info richiesta prima di completarla
    richiesta = db.get_richiesta(richiesta_id)
    
    if richiesta:
        # Completa nel database
        db.complete_richiesta(richiesta_id)
        
        await query.answer("✅ Richiesta completata!")
        
        # Aggiorna messaggio corrente con strikethrough (se message_id disponibile)
        try:
            if richiesta.get('message_id'):
                updated_text = (
                    f"~~✅ COMPLETATO~~\n\n"
                    f"👤 {richiesta['user_nome_completo']}\n"
                    f"🏠 {richiesta['appartamento_nome']}\n"
                    f"📦 ~~{richiesta['descrizione_prodotti']}~~\n"
                    f"📅 {richiesta['data_richiesta']}"
                )
                await context.bot.edit_message_text(
                    chat_id=query.message.chat_id,
                    message_id=richiesta['message_id'],
                    text=updated_text,
                    parse_mode='Markdown'
                )
        except Exception as e:
            logger.error(f"Errore aggiornamento messaggio: {e}")
        
        # Notifica l'utente
        try:
            await context.bot.send_message(
                chat_id=richiesta['telegram_id'],
                text=f"✅ *Prodotti consegnati!*\n\n"
                     f"🏠 {richiesta['appartamento_nome']}\n"
                     f"📦 {richiesta['descrizione_prodotti']}\n\n"
                     f"I prodotti richiesti sono stati consegnati.",
                parse_mode='Markdown'
            )
        except Exception as e:
            logger.error(f"Errore notifica utente: {e}")
        
        logger.info(f"Richiesta {richiesta_id} completata dall'admin")
        
        # Aggiorna lista richieste
        await admin_richieste_prodotti(update, context)
    else:
        await query.answer("❌ Richiesta non trovata")


async def pulisci_richieste_completate(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Elimina richieste completate"""
    query = update.callback_query
    
    if not is_admin(update.effective_user.id):
        try:
            await query.answer("❌ Non autorizzato")
        except BadRequest:
            pass
        return
    
    deleted = db.delete_richieste_completate()
    
    try:
        await query.answer(f"🗑️ {deleted} richieste eliminate")
    except BadRequest:
        pass
    
    # Aggiorna lista
    await admin_richieste_prodotti(update, context)


# ==================== REPORT ORE ====================

async def admin_report_ore(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Menu report ore"""
    query = update.callback_query
    await query.answer()
    
    if not is_admin(update.effective_user.id):
        return
    
    keyboard = [
        [InlineKeyboardButton("📅 Oggi", callback_data="report_oggi")],
        [InlineKeyboardButton("📅 Ieri", callback_data="report_ieri")],
        [InlineKeyboardButton("📅 Questa settimana", callback_data="report_settimana")],
        [InlineKeyboardButton("📅 Questo mese", callback_data="report_mese")],
        [InlineKeyboardButton("« Indietro", callback_data="admin_menu")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    text = "⏰ *REPORT ORE*\n\n"
    text += "Seleziona il periodo:"
    
    await query.edit_message_text(text, reply_markup=reply_markup, parse_mode='Markdown')


async def mostra_report(update: Update, context: ContextTypes.DEFAULT_TYPE, 
                        periodo: str):
    """Mostra report ore per periodo selezionato"""
    query = update.callback_query
    await query.answer()
    
    if not is_admin(update.effective_user.id):
        return
    
    # Determina date
    oggi = datetime.now().date()
    
    if periodo == 'oggi':
        data_inizio = data_fine = oggi
        titolo = f"Oggi ({format_data(oggi)})"
    elif periodo == 'ieri':
        ieri = oggi - timedelta(days=1)
        data_inizio = data_fine = ieri
        titolo = f"Ieri ({format_data(ieri)})"
    elif periodo == 'settimana':
        data_inizio, data_fine = get_settimana_corrente()
        titolo = f"Settimana {format_data(data_inizio)} - {format_data(data_fine)}"
    elif periodo == 'mese':
        data_inizio, data_fine = get_mese_corrente()
        titolo = f"Mese di {data_inizio.strftime('%B %Y')}"
    else:
        return
    
    # Ottieni tutti gli utenti
    utenti = db.get_all_users()
    
    text = f"⏰ *REPORT ORE*\n{titolo}\n\n"
    
    ore_totali_complessive = 0
    
    for user in utenti:
        turni = db.get_turni_by_user(user['telegram_id'], data_inizio, data_fine)
        turni_completati = [t for t in turni if t['status'] == 'completato']
        
        if not turni_completati:
            continue
        
        ore_user = sum(t.get('ore_lavorate', 0) or 0 for t in turni_completati)
        ore_totali_complessive += ore_user
        
        text += f"👤 *{user['nome']} {user['cognome']}*\n"
        
        # Dettaglio per appartamento
        appartamenti = {}
        for t in turni_completati:
            app_nome = t['appartamento_nome']
            if app_nome not in appartamenti:
                appartamenti[app_nome] = 0
            appartamenti[app_nome] += t.get('ore_lavorate', 0) or 0
        
        for app_nome, ore in appartamenti.items():
            text += f"   🏠 {app_nome}: {format_ore(ore)}\n"
        
        text += f"   📊 *Totale: {format_ore(ore_user)}*\n\n"
    
    if ore_totali_complessive == 0:
        text += "_Nessun turno completato in questo periodo._"
    else:
        text += f"━━━━━━━━━━━━━━━━\n"
        text += f"🎯 *TOTALE GENERALE: {format_ore(ore_totali_complessive)}*"
    
    keyboard = [
        [InlineKeyboardButton("📊 Export Excel", callback_data=f"export_{periodo}")],
        [InlineKeyboardButton("« Indietro", callback_data="admin_report")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(text, reply_markup=reply_markup, parse_mode='Markdown')


# ==================== ARCHIVIO VIDEO ====================

async def admin_archivio_video(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Mostra archivio video per data"""
    query = update.callback_query
    await query.answer()
    
    if not is_admin(update.effective_user.id):
        return
    
    keyboard = [
        [InlineKeyboardButton("📅 Oggi", callback_data="video_oggi")],
        [InlineKeyboardButton("📅 Ieri", callback_data="video_ieri")],
        [InlineKeyboardButton("📊 Statistiche storage", callback_data="video_stats")],
        [InlineKeyboardButton("« Indietro", callback_data="admin_menu")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    text = "📹 *ARCHIVIO VIDEO*\n\n"
    text += "Seleziona un'opzione:"
    
    await query.edit_message_text(text, reply_markup=reply_markup, parse_mode='Markdown')


async def mostra_video_per_data(update: Update, context: ContextTypes.DEFAULT_TYPE, giorno: str):
    """Mostra video di una specifica data"""
    query = update.callback_query
    await query.answer()
    
    if not is_admin(update.effective_user.id):
        return
    
    # Determina data
    oggi = datetime.now().date()
    
    if giorno == 'oggi':
        data = oggi
    elif giorno == 'ieri':
        data = oggi - timedelta(days=1)
    else:
        return
    
    # Ottieni turni della data
    turni = db.get_turni_by_date(data)
    
    text = f"📹 *ARCHIVIO VIDEO*\n{format_data_italiana(data)}\n\n"
    
    if not turni:
        text += "_Nessun video disponibile per questa data._"
        keyboard = [[InlineKeyboardButton("« Indietro", callback_data="admin_video")]]
    else:
        text += f"_Trovati {len(turni)} turni_\n\n"
        
        keyboard = []
        
        for turno in turni:
            text += f"🏠 *{turno['appartamento_nome']}*\n"
            text += f"👤 {turno['nome']} {turno['cognome']}\n"
            
            if turno.get('timestamp_ingresso'):
                ts_ing = datetime.fromisoformat(turno['timestamp_ingresso'])
                text += f"   ⏰ Ingresso: {format_ora(ts_ing)}\n"
            
            if turno.get('timestamp_uscita'):
                ts_usc = datetime.fromisoformat(turno['timestamp_uscita'])
                text += f"   ⏰ Uscita: {format_ora(ts_usc)}\n"
                text += f"   ⏱️  {format_ore(turno.get('ore_lavorate', 0))}\n"
            
            text += "\n"
            
            # Bottoni per vedere i video
            buttons_row = []
            if turno.get('video_ingresso_file_id'):
                buttons_row.append(
                    InlineKeyboardButton("▶️ Ingresso", callback_data=f"play_ing_{turno['id']}")
                )
            if turno.get('video_uscita_file_id'):
                buttons_row.append(
                    InlineKeyboardButton("▶️ Uscita", callback_data=f"play_usc_{turno['id']}")
                )
            
            if buttons_row:
                keyboard.append(buttons_row)
        
        keyboard.append([InlineKeyboardButton("« Indietro", callback_data="admin_video")])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(text, reply_markup=reply_markup, parse_mode='Markdown')


async def play_video(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Invia video all'admin"""
    query = update.callback_query
    
    if not is_admin(update.effective_user.id):
        await query.answer("❌ Non autorizzato")
        return
    
    # Parse callback data: play_ing_123 o play_usc_123
    parts = query.data.split('_')
    tipo = 'ingresso' if parts[1] == 'ing' else 'uscita'
    turno_id = int(parts[2])
    
    # Ottieni turno da Excel
    import openpyxl
    from database import EXCEL_TURNI_PATH
    
    turno = None
    try:
        wb = openpyxl.load_workbook(EXCEL_TURNI_PATH, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == turno_id:
                turno = {
                    'nome': row[2],
                    'cognome': row[3],
                    'appartamento_nome': row[6],
                    'timestamp_ingresso': row[8],
                    'timestamp_uscita': row[9],
                    'video_ingresso_file_id': row[12],
                    'video_uscita_file_id': row[14]
                }
                break
        wb.close()
    except:
        pass
    
    if not turno:
        await query.answer("❌ Turno non trovato")
        return
    
    await query.answer(f"📹 Invio video {tipo}...")
    
    # Determina file_id
    if tipo == 'ingresso':
        file_id = turno.get('video_ingresso_file_id')
        timestamp = turno.get('timestamp_ingresso')
    else:
        file_id = turno.get('video_uscita_file_id')
        timestamp = turno.get('timestamp_uscita')
    
    if not file_id:
        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text=f"❌ Video {tipo} non disponibile"
        )
        return
    
    # Crea caption
    ts = datetime.fromisoformat(timestamp) if timestamp else None
    caption = (
        f"📹 Video {tipo}\n"
        f"👤 {turno['nome']} {turno['cognome']}\n"
        f"🏠 {turno['appartamento_nome']}\n"
        f"⏰ {format_ora(ts) if ts else 'N/A'}"
    )
    
    # Invia video
    success = await send_video_by_file_id(
        context, 
        update.effective_chat.id,
        file_id,
        caption
    )
    
    if not success:
        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text="❌ Errore durante l'invio del video"
        )


async def mostra_stats_storage(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Mostra statistiche storage video"""
    query = update.callback_query
    await query.answer()
    
    if not is_admin(update.effective_user.id):
        return
    
    stats = get_storage_stats()
    
    text = "📊 *STATISTICHE STORAGE*\n\n"
    text += f"📹 Video totali: *{stats['total_files']}*\n"
    text += f"💾 Spazio occupato:\n"
    text += f"   • {stats['total_size_mb']:.2f} MB\n"
    text += f"   • {stats['total_size_gb']:.2f} GB\n"
    
    keyboard = [[InlineKeyboardButton("« Indietro", callback_data="admin_video")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(text, reply_markup=reply_markup, parse_mode='Markdown')


# ==================== GESTIONE UTENTI ====================

async def admin_gestione_utenti(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Mostra lista utenti registrati"""
    query = update.callback_query
    await query.answer()
    
    if not is_admin(update.effective_user.id):
        return
    
    utenti = db.get_all_users()
    
    text = f"👥 *GESTIONE UTENTI* ({len(utenti)})\n\n"
    
    for user in utenti:
        text += f"👤 *{user['nome']} {user['cognome']}*\n"
        text += f"   ID: `{user['telegram_id']}`\n"
        
        if user.get('username'):
            text += f"   @{user['username']}\n"
        
        # Statistiche rapide
        ore_totali = db.get_ore_totali_user(user['telegram_id'])
        text += f"   ⏱️  Ore totali: {format_ore(ore_totali)}\n"
        
        text += "\n"
    
    keyboard = [[InlineKeyboardButton("« Indietro", callback_data="admin_menu")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(text, reply_markup=reply_markup, parse_mode='Markdown')


# ==================== STATISTICHE GENERALI ====================

async def admin_statistiche(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Mostra statistiche generali del sistema"""
    query = update.callback_query
    await query.answer()
    
    if not is_admin(update.effective_user.id):
        return
    
    # Conta utenti
    utenti = db.get_all_users()
    num_utenti = len(utenti)
    
    # Conta turni da Excel
    import openpyxl
    from database import EXCEL_TURNI_PATH
    
    turni_completati = 0
    turni_in_corso = 0
    ore_totali = 0.0
    
    try:
        wb = openpyxl.load_workbook(EXCEL_TURNI_PATH, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[15] == 'completato':
                turni_completati += 1
                ore_totali += row[10] or 0
            elif row[15] == 'in_corso':
                turni_in_corso += 1
        wb.close()
    except:
        pass
    
    richieste_pending = len(db.get_richieste_non_completate())
    
    # Storage
    storage = get_storage_stats()
    
    text = "📊 *STATISTICHE SISTEMA*\n\n"
    
    text += "👥 *Utenti*\n"
    text += f"   • Registrati: {num_utenti}\n\n"
    
    text += "🏠 *Turni*\n"
    text += f"   • Completati: {turni_completati}\n"
    text += f"   • In corso: {turni_in_corso}\n"
    text += f"   • Ore totali: {format_ore(ore_totali)}\n\n"
    
    text += "📦 *Richieste*\n"
    text += f"   • In sospeso: {richieste_pending}\n\n"
    
    text += "📹 *Video*\n"
    text += f"   • Totali: {storage['total_files']}\n"
    text += f"   • Spazio: {storage['total_size_gb']:.2f} GB\n"
    
    keyboard = [[InlineKeyboardButton("« Indietro", callback_data="admin_menu")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(text, reply_markup=reply_markup, parse_mode='Markdown')


# ==================== ROUTING CALLBACKS ====================

async def admin_callback_router(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Router per callback admin"""
    query = update.callback_query
    data = query.data
    
    if not is_admin(update.effective_user.id):
        await query.answer("❌ Non autorizzato")
        return
    
    # Menu principale
    if data == "admin_menu":
        await cmd_admin(update, context)
    
    # Turni
    elif data == "admin_turni":
        await admin_turni_in_corso(update, context)
    elif data == "admin_turni_menu":
        await admin_turni_menu(update, context)
    elif data == "admin_turni_finiti":
        await admin_turni_finiti(update, context)
    elif data == "admin_turni_oggi":
        await admin_turni_oggi(update, context)
    elif data == "admin_turni_globali":
        await admin_turni_globali(update, context)
    elif data == "admin_export_turni_oggi":
        await admin_export_turni(update, context, tipo='oggi')
    elif data == "admin_export_turni_globali":
        await admin_export_turni(update, context, tipo='globali')
    
    # Richieste
    elif data == "admin_richieste":
        await admin_richieste_prodotti(update, context)
    elif data.startswith("completa_"):
        await completa_richiesta(update, context)
    elif data == "admin_pulisci_richieste":
        await pulisci_richieste_completate(update, context)
    
    # Report
    elif data == "admin_report":
        await admin_report_ore(update, context)
    elif data.startswith("report_"):
        periodo = data.split('_')[1]
        await mostra_report(update, context, periodo)
    
    # Video
    elif data == "admin_video":
        await admin_archivio_video(update, context)
    elif data.startswith("video_"):
        giorno = data.split('_')[1]
        if giorno == 'stats':
            await mostra_stats_storage(update, context)
        else:
            await mostra_video_per_data(update, context, giorno)
    elif data.startswith("play_"):
        await play_video(update, context)
    
    # Utenti
    elif data == "admin_utenti":
        await admin_gestione_utenti(update, context)
    
    # Statistiche
    elif data == "admin_stats":
        await admin_statistiche(update, context)
