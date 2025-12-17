#!/usr/bin/env python
"""
Script per aggiungere le coordinate GPS agli appartamenti nel database Excel.
Usa Google Maps Geocoding API per convertire gli indirizzi in coordinate.
"""

import os
import sys

# Aggiungi il percorso del progetto
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from funzioni.google_maps_helper import geocode_address
from funzioni.config import GOOGLE_MAPS_API_KEY

# Path del database
DATABASE_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), '..', 'Database', 'appartamenti.xlsx')

def main():
    print("=" * 60)
    print("🗺️  AGGIUNGI COORDINATE GPS AGLI APPARTAMENTI")
    print("=" * 60)
    
    if not GOOGLE_MAPS_API_KEY:
        print("❌ Google Maps API Key non configurata!")
        return
    
    print(f"\n📂 Database: {DATABASE_PATH}")
    
    if not os.path.exists(DATABASE_PATH):
        print(f"❌ File non trovato: {DATABASE_PATH}")
        return
    
    # Carica il workbook
    wb = openpyxl.load_workbook(DATABASE_PATH)
    ws = wb.active
    
    # Verifica se la colonna Coordinate GPS esiste già
    headers = [cell.value for cell in ws[1]]
    print(f"\n📋 Colonne attuali: {headers}")
    
    coord_col = None
    if 'Coordinate GPS' in headers:
        coord_col = headers.index('Coordinate GPS') + 1
        print(f"✅ Colonna 'Coordinate GPS' già presente (colonna {coord_col})")
    else:
        # Aggiungi la colonna alla fine
        coord_col = len(headers) + 1
        ws.cell(row=1, column=coord_col, value='Coordinate GPS')
        print(f"➕ Aggiunta colonna 'Coordinate GPS' (colonna {coord_col})")
    
    # Indice colonne (1-based per openpyxl)
    NOME_COL = 2  # Ciao Booking Nome
    INDIRIZZO_COL = 4  # Indirizzo
    
    # Processa ogni appartamento
    totale = 0
    geocodati = 0
    errori = 0
    saltati = 0
    
    print("\n🔄 Elaborazione appartamenti...\n")
    
    for row_num in range(2, ws.max_row + 1):
        nome = ws.cell(row=row_num, column=NOME_COL).value
        indirizzo = ws.cell(row=row_num, column=INDIRIZZO_COL).value
        coord_esistenti = ws.cell(row=row_num, column=coord_col).value
        
        if not nome:
            continue
        
        totale += 1
        
        # Se ha già coordinate valide, salta
        if coord_esistenti and str(coord_esistenti).strip() and ',' in str(coord_esistenti):
            print(f"  ⏭️  {nome}: già geocodato ({coord_esistenti})")
            saltati += 1
            continue
        
        if not indirizzo or str(indirizzo).strip() == '':
            print(f"  ⚠️  {nome}: indirizzo mancante")
            errori += 1
            continue
        
        # Geocoding
        try:
            coords = geocode_address(str(indirizzo))
            if coords:
                lat, lon = coords
                coord_str = f"{lat},{lon}"
                ws.cell(row=row_num, column=coord_col, value=coord_str)
                print(f"  ✅ {nome}: {coord_str}")
                geocodati += 1
            else:
                print(f"  ❌ {nome}: geocoding fallito per '{indirizzo}'")
                errori += 1
        except Exception as e:
            print(f"  ❌ {nome}: errore - {e}")
            errori += 1
    
    # Salva
    print("\n💾 Salvataggio...")
    wb.save(DATABASE_PATH)
    wb.close()
    
    print("\n" + "=" * 60)
    print(f"📊 RIEPILOGO:")
    print(f"   • Totale appartamenti: {totale}")
    print(f"   • Geocodati ora: {geocodati}")
    print(f"   • Già geocodati: {saltati}")
    print(f"   • Errori: {errori}")
    print("=" * 60)
    print("✅ Completato!")

if __name__ == "__main__":
    main()
